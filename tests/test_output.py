from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook
from pyrsistencesniper.models.finding import AccessLevel, Enrichment, Finding
from pyrsistencesniper.output.csv_output import CsvOutput, _sanitize_cell
from pyrsistencesniper.output.html_output import HtmlOutput
from pyrsistencesniper.output.xlsx_output import XlsxOutput

# -- CSV formula injection tests -----------------------------------------------


def test_sanitize_cell_equals() -> None:
    assert _sanitize_cell("=CMD()") == "'=CMD()"


def test_sanitize_cell_plus() -> None:
    assert _sanitize_cell("+1+2") == "'+1+2"


def test_sanitize_cell_minus() -> None:
    assert _sanitize_cell("-1-2") == "'-1-2"


def test_sanitize_cell_at() -> None:
    assert _sanitize_cell("@SUM(A1)") == "'@SUM(A1)"


def test_sanitize_cell_tab() -> None:
    assert _sanitize_cell("\t=CMD()") == "'\t=CMD()"


def test_sanitize_cell_carriage_return() -> None:
    assert _sanitize_cell("\r=CMD()") == "'\r=CMD()"


def test_sanitize_cell_newline() -> None:
    assert _sanitize_cell("\n=CMD()") == "'\n=CMD()"


def test_sanitize_cell_whitespace_prefix() -> None:
    """Leading whitespace before a formula character should be caught."""
    assert _sanitize_cell(" =CMD()") == "' =CMD()"


def test_sanitize_cell_safe_value() -> None:
    assert _sanitize_cell("explorer.exe") == "explorer.exe"


def test_sanitize_cell_empty() -> None:
    assert _sanitize_cell("") == ""


def test_sanitize_cell_number() -> None:
    assert _sanitize_cell(42) == "42"


# -- HTML autoescaping tests ---------------------------------------------------


def _make_result(
    path: str = "HKLM\\Run", value: str = "test.exe"
) -> tuple[Finding, tuple[Enrichment, ...]]:
    finding = Finding(
        path=path,
        value=value,
        technique="Test",
        mitre_id="T0000",
        description="Test description",
        access_gained=AccessLevel.SYSTEM,
        hostname="HOST",
        check_id="test_check",
    )
    return (finding, ())


def test_html_autoescaping_value() -> None:
    """A <script> tag in value must be escaped in HTML output."""
    result = _make_result(value="<script>alert(1)</script>")
    out = io.StringIO()
    renderer = HtmlOutput()
    renderer._write([result], out)
    html = out.getvalue()
    assert "<script>" not in html
    assert "&lt;script&gt;" in html


def test_html_autoescaping_path() -> None:
    """A <script> tag in path must be escaped in HTML output."""
    result = _make_result(path='HKLM\\<img src=x onerror="alert(1)">')
    out = io.StringIO()
    renderer = HtmlOutput()
    renderer._write([result], out)
    html = out.getvalue()
    assert 'onerror="alert(1)"' not in html
    assert "&lt;img" in html


def test_csv_output_sanitizes_all_fields() -> None:
    """All cell values in CSV output should be sanitized."""
    result = _make_result(value="=HYPERLINK()")
    out = io.StringIO()
    renderer = CsvOutput()
    renderer._write([result], out)
    csv_text = out.getvalue()
    assert "'=HYPERLINK()" in csv_text


# -- XLSX output tests ---------------------------------------------------------


def test_xlsx_produces_valid_workbook() -> None:
    """render() should produce a valid XLSX with correct headers and data."""
    result = _make_result(value="malware.exe")
    buf = io.BytesIO()
    XlsxOutput().render([result], output=buf)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "path" in headers
    assert "value" in headers
    assert ws.cell(row=2, column=headers.index("value") + 1).value == "malware.exe"


def test_xlsx_raises_without_output() -> None:
    """render() must raise ValueError when output is None."""
    with pytest.raises(ValueError, match="requires a file path"):
        XlsxOutput().render([_make_result()])


def test_xlsx_enrichment_columns() -> None:
    """Dynamic enrichment columns should appear in the header."""
    finding = Finding(
        path="HKLM\\Run",
        value="test.exe",
        technique="Test",
        mitre_id="T0000",
        description="d",
        access_gained=AccessLevel.SYSTEM,
        hostname="HOST",
        check_id="test_check",
    )
    enrichment = Enrichment(provider="vt", data={"score": "5/70"})
    result = (finding, (enrichment,))
    buf = io.BytesIO()
    XlsxOutput().render([result], output=buf)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "enrichment.vt.score" in headers
    assert (
        ws.cell(row=2, column=headers.index("enrichment.vt.score") + 1).value == "5/70"
    )


def test_xlsx_empty_results() -> None:
    """Empty results should produce a workbook with only a header row."""
    buf = io.BytesIO()
    XlsxOutput().render([], output=buf)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "path"
    assert ws.cell(row=2, column=1).value is None
