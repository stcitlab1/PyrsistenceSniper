from __future__ import annotations

from pathlib import Path
from typing import IO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from pyrsistencesniper.models.finding import AnnotatedResult
from pyrsistencesniper.output.base import OutputBase

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


class XlsxOutput(OutputBase):
    """Writes findings as a styled XLSX workbook."""

    def render(
        self,
        results: list[AnnotatedResult],
        output: Path | IO[bytes] | None = None,  # type: ignore[override]
    ) -> None:
        if output is None:
            msg = "XLSX output requires a file path or binary stream"
            raise ValueError(msg)

        rows, fieldnames = self._flatten_results(results)

        wb = Workbook()
        ws = wb.active
        ws.title = "Findings"

        # Header row
        for col_idx, name in enumerate(fieldnames, start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, name in enumerate(fieldnames, start=1):
                ws.cell(row=row_idx, column=col_idx, value=str(row.get(name, "")))

        # Auto-fit column widths (approximate)
        for col_idx, name in enumerate(fieldnames, start=1):
            max_len = len(name)
            for row_idx in range(2, len(rows) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = min(max_len + 2, 60)

        # Freeze header row
        ws.freeze_panes = "A2"

        if isinstance(output, Path):
            wb.save(str(output))
        else:
            wb.save(output)

    def _write(self, results: list[AnnotatedResult], out: IO[str]) -> None:
        msg = "XLSX is binary; use render() directly"
        raise NotImplementedError(msg)
